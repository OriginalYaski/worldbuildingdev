#Generates an entire continent's worth of characters all at once
#Outdated

import random
from openpyxl import Workbook
from openpyxl import load_workbook

#for the progress display
CURRENT_PERCENT = 0

#------------------------------------------------------------------------

#choose a random level of PC
def find_PC(current, roll, level, pc_list):

#start from level 1 and go upwards
    level_pointer = 0
#loop through levels
    while roll > 0:
        level_pointer += 1
#at each level, offset the roll by the number of PCs remaining in that level
#when it crosses the threshold, exit loop
        roll -= pc_list[20 - level_pointer]

    """
    key = "PC: lvl " + str(level_pointer)
    if key in level:
        level[key] += 1
    else:
        level[key] = 1
        """
#choose a random PC from the chosen level
    level = grab_PC(level_pointer, roll, level, pc_list)
    
    return level

#------------------------------------------------------------------------

#choose a random PC from a given level
def grab_PC(current, roll, level, total):
#iterate through the various classes of that level
    for cell_pointer in range(0,23):
#add the number of PCs of that class to the current score
        roll += level_array[current][cell_pointer]
#if threshold is passed, choose one of the PCs of that class
        if roll > 0:
#increment the totals to indicate a PC of this class has been selected
            level_array[current][cell_pointer] -= 1
            total[20 - current] -= 1
#grab the PC class name to create the dictionary entry
            key = PC_class_list[cell_pointer] + ": lvl " + str(current)
#update the dictionary
            if key in level:
                level[key] += 1
            else:
                level[key] = 1

            break
#return the dictionary    
    return level

#------------------------------------------------------------------------

#method to generate 1 level of the dictionary
def generate_level(current, NPCs, PCs, total, tracker):
#generate the sub-dictionary
    level_dict = {"NPCs":0}

    PCs += tracker[20-current]

    Adults = NPCs + PCs

#run until the level is full
    while total > 0:
#choose a random individual from amongst all the remaining adults
        r = random.randrange(1, Adults + 1)
#if the number is greater than the total remaining PCs, it must be an NPC
        if r > PCs:
            level_dict["NPCs"] += 1
            NPCs -= 1
        else:
#if the number is less than or equal to the remaining number of PCs, find that PC and add them
            level_dict = find_PC(current, r, level_dict, tracker)

            PCs -= 1
#update the number of remaining adults of that level, and overall
        total -= 1
        Adults -=1

#should probably have either progress display or level display, not both.
        """
        if Adults % 6400000 == 0:
            global CURRENT_PERCENT
            CURRENT_PERCENT += 1
            print(str(CURRENT_PERCENT) + "% done...")
            """
        
    print("Level " + str(current) + ":")
    print(level_dict)
    return NPCs, PCs, level_dict

#------------------------------------------------------------------------

#function to generate levels 1-20 as a dictionary
def dict_make(NPCs, ws, class_list):

#Grab the per level sum totals of the PC and general populations
    PC_tracker = [ws.cell(row = i, column = 24).value for i in range(1,21)]
    total_tracker = [ws.cell(row = i, column = 25).value for i in range(1,21)]

#instantiate the dictionary
    demographics = {}
#instantiate the list pointer
    x = 19
    PCs = 0
    current = 1

    while x >= 0:
        #loop through each level
        NPCs, PCs, demographics[current] = generate_level(current, NPCs, PCs,
                                               total_tracker[x], PC_tracker)

#increment the pointers
        x -= 1
        current += 1
#return results
    return demographics
    
#------------------------------------------------------------------------

#todo make it accept variable user input
wb = load_workbook('C:/Users/infin/OneDrive/Documents/The big organizer.xlsx')

ws = wb['Noromar big sheet 1']

#get the list of class priority
PC_class_list = [ws.cell(row = 21, column = i).value for i in range(1,24)]
#acquire the class distribution array
level_array = {}
level_key = 15

#iterate through the array
for row in ws.iter_rows(min_row = 6, max_row = 20, max_col = 23, values_only=True):
    #row default returns a tuple, so convert to list for future manipulation
    level_array[level_key] = list(row)
    level_key -=1


"""
#create a flag for when we've found an acceptable output
isgood = 0

while isgood == 0:

    #levels 20-16 can be generated quickly, and have the highest variance
    #offer a breakpoint to see if those levels are satisfactory, and reroll them if desired
    outputlist = the_first_five(ws['AA2'], ws['AA1'], 20, ws, PC_class_list)

    #print(outputlist[5])

    response = input('Is this acceptable? (Y/N):')

#filter out unrecognized commands
    while response != 'Y' and response != 'N':
        response = input('Response not recognized, please enter Y or N:')

    if response == 'Y':
        print('Continuing...\n')
        isgood = 1
    else:
        print('Rerolling...\n')

#receive the outputs from after the first 5 levels are calculated
Adult_total = outputlist[0]
PC_total = outputlist[1]
Current_level = outputlist[2]
PC_level_tracker =outputlist[3]
total_level_tracker = outputlist[4]
total_demographics = outputlist[5]"""

total_demographics = dict_make(ws['AA3'].value, ws, PC_class_list)

#to do, make it output automatically to either an excel sheet or a txt file
#for x in total_demographics.keys():
#    print(total_demographics[x])

wb.close()
