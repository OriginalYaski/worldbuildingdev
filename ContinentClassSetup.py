#Generates an entire distribution of PCs and NPCs for a single continent
#Generates levels 20-16 first, as the top 5 levels most fully describe the feel of the continent
#This feature outdates ClassBuildUp

import random
import copy
from openpyxl import Workbook
from openpyxl import load_workbook

#for the progress display
#CURRENT_PERCENT = 0

#------------------------------------------------------------------------

#choose a random level of PC
def find_PC(PCs, current, roll, level, pc_list):

#start from level 1 and go upwards
    level_pointer = 0
#loop through levels
    while roll > 0:
        level_pointer += 1
#at each level, offset the roll by the number of PCs remaining in that level
#when it crosses the threshold, exit loop
        roll -= pc_list[20 - level_pointer]

    """
    key = "PC: lvl " + str(level_pointer)
    if key in level:
        level[key] += 1
    else:
        level[key] = 1
        """
#choose a random PC from the chosen level
    level = grab_PC(level_pointer, roll, level, pc_list)
    
    return level

#------------------------------------------------------------------------

#choose a random PC from a given level
def grab_PC(current, roll, level, total):
#iterate through the various classes of that level
    for cell_pointer in range(0,mark-1):
#add the number of PCs of that class to the current score
        if level_array[current][cell_pointer] is not None:
            roll += level_array[current][cell_pointer]
#if threshold is passed, choose one of the PCs of that class
        if roll > 0:
#increment the totals to indicate a PC of this class has been selected
            level_array[current][cell_pointer] -= 1
            total[20 - current] -= 1
#update the dictionary
            if current in level:
                level[current][cell_pointer] += 1
            else:
                level[current] = [0]*(mark-1)
                level[current][cell_pointer] = 1

            break
#return the dictionary    
    return level

#------------------------------------------------------------------------

#method to generate 1 level of the dictionary
def generate_level(current, Adults, PCs, total, tracker):
#generate the sub-dictionary
    level_dict = {current:[0]*mark}

#run until the level is full
    while total > 0:
#if the number of PCs of that level matches the remaining number of adults of that level
#then all the adults of that level must be the remaining PCs
        if tracker[20-current] == total:
            level_dict = grab_PC(current, 0, level_dict, tracker)

            PCs -= 1

        else:
#choose a random individual from amongst all the remaining adults
            r = random.randrange(1, Adults + 1)
#if the number is greater than the total remaining PCs, it must be an NPC
            if r > PCs:
                level_dict[current][mark-1] += 1
            else:
#if the number is less than or equal to the remaining number of PCs, find that PC and add them
                level_dict = find_PC(PCs, current, r, level_dict, tracker)

                PCs -= 1
#update the number of remaining adults of that level, and overall
        total -= 1
        Adults -=1

#should probably have either progress display or level display, not both.
        """
        if Adults % 6400000 == 0:
            global CURRENT_PERCENT
            CURRENT_PERCENT += 1
            print(str(CURRENT_PERCENT) + "% done...")
            """

#Reformat output into a manner that is much shorter and more legible than printing the entire dict
    print("\nLevel " + str(current) + ":")
#We'll always have NPCs, and if we don't that's notable too
    print("NPCs: " + str(level_dict[current][mark-1]))
    
    for cl in range(0,mark-1):
#Setup to print a Class type
        printable = PC_class_list[cl]
        isthere = False

#Search for if there are any PCs of that level
        for lev in range(20,0,-1):
            if lev in level_dict:

#Check if any of the PCs of that level are of the right class
                if level_dict[lev][cl] != 0:
#Mark the class as present
                    isthere = True

#Add that level to the output string
                    printable += ": " + str(lev) + "(" + str(level_dict[lev][cl]) + ") "

#If there are any PCs of that class, let us know
        if isthere:
            print(printable)
    
    #print(level_dict)
    return Adults, PCs, level_dict

#------------------------------------------------------------------------

#function to generate levels 20-11 as a dictionary
def the_first_ten(Adults, PCs, current, ws):

#Grab the per level sum totals of the PC and general populations
    PC_tracker = [ws.cell(row = i, column = mark+1).value for i in range(1,21)]
    total_tracker = [ws.cell(row = i, column = mark+2).value for i in range(20,0,-1)]

#instantiate the dictionary
    demographics = {}

    while current > 10:
        #loop through each level
        Adults, PCs, demographics[current] = generate_level(current, Adults, PCs,
                                               total_tracker[current-1], PC_tracker)

#increment the pointer
        current -= 1
#return results
    return [Adults, PCs, current, PC_tracker, total_tracker, demographics]

#------------------------------------------------------------------------

#function to generate levels 15-1 and add them to the dictionary
def the_rest(Adults, PCs, current, PC_tracker, total_tracker, demographic):
    
    while current > 1:
        #loop through each level
        Adults, PCs, demographic[current] = generate_level(current, Adults, PCs,
                                               total_tracker[current-1], PC_tracker)

#increment the pointer
        current -= 1

#Skip level 1, it's easy to figure out, but take's too long through generate_level
    demographic[current] = level_array[current] + [Adults - PCs]

#Print it
    print("\nLvl 1:")
    
    for cl in range(0,mark):
        print(PC_class_list[cl] + ": " + str(demographic[1][cl]))
#return finished dictionary
    return demographic
    
#------------------------------------------------------------------------

#todo make it accept variable user input
wb = load_workbook('The new organizer.xlsx')

ws1 = wb['Big sheet 1']

#Get the endpoint of the list
global mark
mark = int(ws1['A1'].value + 1)

#get the list of class priority
PC_class_list = [ws1.cell(row = 21, column = i).value for i in range(2,mark+1)]
PC_class_list += ["NPCs"]
print(PC_class_list)

#Get the totals
PC_total = int(ws1.cell(row = 1, column = mark + 4).value)
Adult_total = int(ws1.cell(row = 2, column = mark + 4).value)

#acquire the class distribution array
master_level_array = {}
level_key = 20

#iterate through the array
for row in ws1.iter_rows(min_row = 1, max_row = 20,
                         min_col = 2, max_col = mark, values_only=True):
    #row default returns a tuple, so convert to list for future manipulation
    master_level_array[level_key] = list(row)
    level_key -=1

#create a flag for when we've found an acceptable output
isgood = 0

while isgood == 0:
#Create dummy values to pass, so the globals don't get modified
    PCs = PC_total
    Adults = Adult_total
    level_array = copy.deepcopy(master_level_array)
    

    #levels 20-11 can be generated quickly, and have the highest variance
    #offer a breakpoint to see if those levels are satisfactory, and reroll them if desired
    outputlist = the_first_ten(Adults, PCs, 20, ws1)

    #print(outputlist[0])

    response = input('Is this acceptable? (Y/N):')

#filter out unrecognized commands
    while response != 'Y' and response != 'N':
        response = input('Response not recognized, please enter Y or N:')

    if response == 'Y':
        print('Continuing...\n')
        isgood = 1
    else:
        print('Rerolling...\n')

#receive the outputs from after the first 5 levels are calculated
Adult_total = outputlist[0]
PC_total = outputlist[1]
Current_level = outputlist[2]
PC_level_tracker =outputlist[3]
total_level_tracker = outputlist[4]
total_demographics = outputlist[5]

total_demographics = the_rest(Adult_total, PC_total, Current_level, PC_level_tracker,
                               total_level_tracker, total_demographics)

#to do, make it output automatically to either an excel sheet or a txt file
#for x in total_demographics.keys():
#    print(total_demographics[x])
ws2 = wb.create_sheet('Big sheet 2')

wb.save('The new organizer.xlsx')
wb.close()

txt_exist = False
levels = False
try:
    text = open("The new organizer.txt", "r")

    txt_exist = True
    fields = text.readline()

    fields = fields.split(',')

    if "Level Dict" in fields:
        levels = True
finally:
    text.close()


if txt_exist:
    text = open("The new organizer.txt", "r")
    full_file = text.readlines()
    text.close()
else:
    full_file = []

text = open("The new organizer.txt", "w")
text.close()
