#Step one of getting the baseline level distribution (without extra NPC levels)

import random
from openpyxl import Workbook
from openpyxl import load_workbook

#for the progress display
CURRENT_PERCENT = 0

#------------------------------------------------------------------------

#method to setup the PC and total per-level values
def total_lists(sheet, total, col, falloff, scale):
    #Account for potential upper remainders
    tracker = total

    #Create output list
    output = [0,]*20

    for i in range(0, 19):
        #Get value for the level
        val = round(total/(falloff**(i/scale))-total/(falloff**((i+1)/scale)))

        #Write value to cell
        ws.cell(row = 20 - i, column = col, value = val)

        #Add value to list
        output[i] = val

        #print(val)

        #Increment tracker
        tracker -= val

    #We aren't going to epic levels, so keep all the rest
    ws.cell(row = 1, column = col, value = tracker)
    output[19] = tracker
    #print(tracker)

    return output

#------------------------------------------------------------------------

#method to generate 1 level of the dictionary
def generate_level(current, total, total_list, PCs):
#generate the level list
    level_list = [0,]*(mark-1)

#run until the level is full
    while PCs > 0:
#choose a random individual from amongst all the remaining PCs
        r = random.randrange(1, total + 1)

#find the correct PC
        list_pointer = -1

        while r > 0:
            list_pointer += 1
            r -= total_list[list_pointer]

#Add PC to that level and remove from general pool
        total_list[list_pointer] -= 1
        level_list[list_pointer] += 1

#update the number of remaining PCs of that level, and overall
        PCs -= 1
        total -= 1

#should probably have either progress display or level display, not both.
        """
        if Adults % 6400000 == 0:
            global CURRENT_PERCENT
            CURRENT_PERCENT += 1
            print(str(CURRENT_PERCENT) + "% done...")
            """
#Generate a displayable string for user reference        
    level_string = ""

    for i in range(0, mark-1):
        if level_list[i] != 0:
            level_string += PC_class_list[i] + ": " + str(level_list[i]) + ", "

    print("Level " + str(current) + ":")
    print(level_string)
    return total, level_list

#------------------------------------------------------------------------

#function to generate levels 20-16 as a dictionary
def the_first_ten():

#get the list of PCs of each class and refresh the total
#Have to do this here, because otherwise the global variable gets changed each reroll
    total = ws.cell(row = 1, column = mark + 4).value
    PC_total_list = [ws.cell(row = 22, column = i).value for i in range(2,mark+1)]

    s = 0
    for i in range(0, mark-1):
        s += PC_total_list[i]

    print(s)
#instantiate the dictionary
    demographics = {}
#instantiate the list pointer
    current = 20

    while current > 10:
        #loop through each level
        total, demographics[current] = generate_level(current, total, PC_total_list,
                                               PC_lvl_list[current-1])

#increment the pointers
        current -= 1
#return results
    return [total, PC_total_list, demographics]

#------------------------------------------------------------------------

#function to generate levels 10-1 and add them to the dictionary
def the_rest(total, PC_total_list, demographics):

    current = 10

    while current > 0:
        #loop through each level
        total, demographics[current] = generate_level(current, total, PC_total_list,
                                               PC_lvl_list[current-1])

#increment the pointer
        current -= 1
#return finished dictionary
    return demographics
    
#------------------------------------------------------------------------

#todo make it accept variable user input
wb = load_workbook('The new organizer.xlsx')

ws = wb['Big sheet 1']

#Get the enpoint of the list
global mark
mark = ws['A1'].value + 1

#get the list of class priority
PC_class_list = [ws.cell(row = 21, column = i).value for i in range(2,mark+1)]

#Get the totals
PC_total = ws.cell(row = 1, column = mark + 4).value
Adult_total = ws.cell(row = 2, column = mark + 4).value

#Setup the PC and total level distributions
PC_lvl_list = total_lists(ws, PC_total, mark+1, 169, 5)
Total_lvl_list = total_lists(ws, Adult_total, mark+2, 169, 5)

#create a flag for when we've found an acceptable output
isgood = 0

while isgood == 0:

    #levels 20-16 can be generated quickly, and have the highest variance
    #offer a breakpoint to see if those levels are satisfactory, and reroll them if desired
    outputlist = the_first_ten()

    #print(outputlist[5])
    print(outputlist[0])

    response = input('Is this acceptable? (Y/N):')

#filter out unrecognized commands
    while response != 'Y' and response != 'N':
        response = input('Response not recognized, please enter Y or N:')

    if response == 'Y':
        print('Continuing...\n')
        isgood = 1
    else:
        print('Rerolling...\n')

#receive the outputs from after the first 10 levels are calculated
PC_total = outputlist[0]
total_list = outputlist[1]
total_demographics = outputlist[2]

total_demographics = the_rest(PC_total, total_list, total_demographics)

#Save the results to the excel doc
for r in range(1,21):
    for c in range(0,mark-1):
        ws.cell(row = 21-r, column = c+2, value = total_demographics[r][c])

wb.save('The new organizer.xlsx')
wb.close()
