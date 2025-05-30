#This file is for the purpose of generating a random number of champions of each class for the appropriate distribution of a single continent
#In the 1.0 build, the order of the classes is fixed, and it can't be rearranged or have new classes added in
#This is the first change desired in future iterations

from openpyxl import Workbook
import random
import math

#------------------------------------------------------------------------

def find_class(class_list, rand_list, r):
    for i in range(len(rand_list)-1, -1, -1):
        if (r <= rand_list[i]):
            class_list[i] += 1
            return class_list

#------------------------------------------------------------------------

loop = 1

while loop == 1:
    answer = str(input("Read from File? (R) or Write own list? (W)\n"))
    if answer == "R" or answer == "W":
        break
    print("Answer not recognized. ")

if answer == "W":
    length = int(input("How long is the class list?\n"))
    while length < 1:
        length = int(input("You must input at least 1 class.\n"))

    print("List classes in order of most to least common")

    randlist = [0] * length
    classes = [0] * length

    classlist = ["blank"] * length

    for i in range(length):
        classlist[i] = str(input("..."))
else:
    length = 23
    randlist = [0] * length
    classes = [0] * length

    classlist = ["Wizard", "Magus", "Alchemist", "Bard", "Investigator",
                 "Summoner", "Sorcerer", "Thaumaturge", "Rogue", "Fighter", "Witch",
                 "Swashbuckler", "Ranger", "Barbarian", "Kineticist", "Druid", "Champion",
                 "Oracle", "Cleric", "Gunslinger", "Inventor", "Psychic", "Monk"]

wb = Workbook()
ws = wb.create_sheet('Big sheet 1')

L = length - 1
randlist[L] = 3 ** (L)

for i in range(L-1, -1, -1):
    randlist[i] = int(randlist[i+1] * 4/3)

for i in range(L-1, -1, -1):
    randlist[i] = randlist[i] + randlist[i+1]

#Used for calibration/confirmation of proper randlist setup (Optional)
#print(randlist)
#print(randlist[length-1]+1)



answer = str(input("Use default setup?"))
if answer == "Y":
    x = 5120000
    y = 640000000
#This needs to be fleshed out
else:
    x = int(input("Number of iterations"))
    #Need to give controls on ratio of adults to exalted, etc., but I'm hardcoding for now
    y = 640000000

#Store the total PCs, total Adults, and total NPCs (respectively)
ws['AA1'].value = x
ws['AA2'].value = y
ws['AA3'].value = y-x

#Generate all the PCs by class, randomly
for i in range(x):
    r = random.randrange(1,randlist[0]+1)

    classes = find_class(classes, randlist, r)
    
#Print and store all the PCs by class
for n in range(len(classes)):
    ws.cell(row = 21, column = n+1, value = classlist[n])
    ws.cell(row = 22, column = n+1, value = classes[n])
    print(classlist[n],": ",classes[n], sep = "")

wb.save('The new organizer.xlsx')
